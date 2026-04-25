import sys
import math
import socket
import argparse
from pathlib import Path
from typing import Union, List

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


def resource_path(relative_path: str) -> Path:
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / relative_path
    return Path(__file__).resolve().parent / relative_path


def get_worksheet(excel_file: str, sheet_name: Union[int, str]):
    path = Path(excel_file)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    wb = load_workbook(filename=excel_file, data_only=True)

    if isinstance(sheet_name, int):
        sheet_names = wb.sheetnames
        if sheet_name < 0 or sheet_name >= len(sheet_names):
            raise IndexError(f"sheet index {sheet_name} is out of range. Sheets: {sheet_names}")
        ws = wb[sheet_names[sheet_name]]
    else:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        ws = wb[sheet_name]

    return ws


def is_blank_separator(value) -> bool:
    if value is None:
        return True
    return str(value).strip() == ""


def find_column_index_by_header(ws, column_name: str) -> int:
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        raise KeyError(f"Column '{column_name}' not found. Available headers: {headers}")
    return headers.index(column_name) + 1


def get_cell_value_by_header(ws, excel_row: int, column_name: str):
    col_idx = find_column_index_by_header(ws, column_name)
    return ws.cell(row=excel_row, column=col_idx).value


def normalize_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def build_payload(cell_value) -> bytes:
    if cell_value is None:
        raise ValueError("Data Matrix cell is empty (None)")

    text = str(cell_value)
    if text == "":
        raise ValueError("Data Matrix cell is empty (empty string)")

    text = text.replace("_x001D_", "\x1d")
    text = text.replace("_x001d_", "\x1d")

    try:
        payload = text.encode("latin-1")
    except UnicodeEncodeError as e:
        raise ValueError(
            "Data Matrix value contains characters outside latin-1. "
            "Please normalize the source data."
        ) from e

    return payload


def extract_ai_21(payload: bytes, max_len: int = 13) -> str:
    try:
        text = payload.decode("latin-1")
    except Exception:
        return ""

    idx = text.find("21")
    if idx == -1:
        return ""

    rest = text[idx + 2:]
    gs_pos = rest.find("\x1d")
    if gs_pos != -1:
        value = rest[:gs_pos]
    else:
        value = rest

    value = value[:max_len].strip()
    if not value:
        return ""

    return f"(21) {value}"


def wrap_text_by_pixels(
    text: str,
    font: ImageFont.FreeTypeFont,
    max_width_px: int,
    max_lines: int,
    draw: ImageDraw.ImageDraw,
) -> List[str]:
    text = normalize_text(text)
    if not text:
        return []

    words = text.split(" ")
    lines: List[str] = []
    current = ""

    def text_width(s: str) -> int:
        if not s:
            return 0
        bbox = draw.textbbox((0, 0), s, font=font)
        return bbox[2] - bbox[0]

    for word in words:
        if text_width(word) <= max_width_px:
            candidate = word if not current else current + " " + word
            if text_width(candidate) <= max_width_px:
                current = candidate
            else:
                if current:
                    lines.append(current)
                    if len(lines) >= max_lines:
                        return lines
                current = word
        else:
            if current:
                lines.append(current)
                if len(lines) >= max_lines:
                    return lines
                current = ""

            chunk = ""
            for ch in word:
                candidate = chunk + ch
                if text_width(candidate) <= max_width_px:
                    chunk = candidate
                else:
                    if chunk:
                        lines.append(chunk)
                        if len(lines) >= max_lines:
                            return lines
                    chunk = ch
            if chunk:
                current = chunk

    if current and len(lines) < max_lines:
        lines.append(current)

    return lines[:max_lines]


def load_font(font_path: str, font_size: int) -> ImageFont.FreeTypeFont:
    path = Path(font_path)
    if not path.exists():
        raise FileNotFoundError(f"Font file not found: {font_path}")
    return ImageFont.truetype(str(path), font_size)


def load_and_prepare_logo(logo_path: str, logo_height_px: int) -> Image.Image:
    external = Path(logo_path)
    if external.exists():
        path = external
    else:
        path = resource_path(logo_path)

    if not path.exists():
        raise FileNotFoundError(f"EAC logo file not found: {path}")

    logo = Image.open(path).convert("RGBA")
    w, h = logo.size
    if h <= 0:
        raise ValueError("Invalid EAC logo size")

    scale = logo_height_px / h
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    logo = logo.resize((new_w, new_h), Image.LANCZOS)

    bg = Image.new("RGBA", logo.size, (255, 255, 255, 255))
    bg.alpha_composite(logo)
    logo = bg.convert("L")
    logo = logo.point(lambda p: 0 if p < 200 else 255, mode="1")
    return logo


def build_text_lines(
    gtin_text: str,
    name_text: str,
    ai21_text: str,
    font_path: str,
    gtin_font_size: int,
    name_font_size: int,
    ai21_font_size: int,
    text_width_px: int,
    name_max_lines: int,
) -> tuple[List[str], List[str], List[str]]:
    gtin_text = normalize_text(gtin_text)
    name_text = normalize_text(name_text)
    ai21_text = normalize_text(ai21_text)

    temp_img = Image.new("L", (text_width_px, 1000), 255)
    draw = ImageDraw.Draw(temp_img)

    gtin_font = load_font(font_path, gtin_font_size)
    name_font = load_font(font_path, name_font_size)
    ai21_font = load_font(font_path, ai21_font_size)

    gtin_lines: List[str] = []
    name_lines: List[str] = []
    ai21_lines: List[str] = []

    if gtin_text:
        gtin_lines = wrap_text_by_pixels(gtin_text, gtin_font, max(40, text_width_px - 12), 1, draw)

    if name_text:
        name_lines = wrap_text_by_pixels(name_text, name_font, max(40, text_width_px - 12), name_max_lines, draw)

    if ai21_text:
        ai21_lines = wrap_text_by_pixels(ai21_text, ai21_font, max(40, text_width_px - 12), 1, draw)

    return gtin_lines, name_lines, ai21_lines


def render_text_block_to_image(
    gtin_lines: List[str],
    name_lines: List[str],
    ai21_lines: List[str],
    font_path: str,
    gtin_font_size: int,
    name_font_size: int,
    ai21_font_size: int,
    width_px: int,
    padding: int,
    line_gap: int,
) -> Image.Image:
    gtin_font = load_font(font_path, gtin_font_size)
    name_font = load_font(font_path, name_font_size)
    ai21_font = load_font(font_path, ai21_font_size)

    tmp = Image.new("L", (width_px, 2000), 255)
    draw = ImageDraw.Draw(tmp)

    items = []
    for line in gtin_lines:
        bbox = draw.textbbox((0, 0), line, font=gtin_font)
        items.append((line, gtin_font, bbox[3] - bbox[1]))
    for line in name_lines:
        bbox = draw.textbbox((0, 0), line, font=name_font)
        items.append((line, name_font, bbox[3] - bbox[1]))
    for line in ai21_lines:
        bbox = draw.textbbox((0, 0), line, font=ai21_font)
        items.append((line, ai21_font, bbox[3] - bbox[1]))

    if items:
        text_height = sum(h for _, _, h in items) + line_gap * (len(items) - 1)
    else:
        text_height = max(gtin_font_size, name_font_size, ai21_font_size)

    height_px = padding * 2 + text_height
    img = Image.new("L", (width_px, height_px), 255)
    draw = ImageDraw.Draw(img)

    y = padding
    for line, font, h in items:
        draw.text((padding, y), line, font=font, fill=0)
        y += h + line_gap

    return img.point(lambda p: 0 if p < 200 else 255, mode="1")


def image_to_zpl_gfa(img: Image.Image) -> bytes:
    if img.mode != "1":
        img = img.convert("1")

    width, height = img.size
    bytes_per_row = math.ceil(width / 8)
    total_bytes = bytes_per_row * height

    pixels = img.load()
    rows_hex = []

    for y in range(height):
        row_bytes = bytearray()
        current_byte = 0
        bit_count = 0

        for x in range(width):
            bit = 1 if pixels[x, y] == 0 else 0
            current_byte = (current_byte << 1) | bit
            bit_count += 1
            if bit_count == 8:
                row_bytes.append(current_byte)
                current_byte = 0
                bit_count = 0

        if bit_count > 0:
            current_byte = current_byte << (8 - bit_count)
            row_bytes.append(current_byte)

        rows_hex.append("".join(f"{b:02X}" for b in row_bytes))

    data = "".join(rows_hex)
    gfa = f"^GFA,{total_bytes},{total_bytes},{bytes_per_row},{data}"
    return gfa.encode("ascii")


def build_zpl(
    payload: bytes,
    text_img: Image.Image,
    eac_img: Image.Image,
    x: int,
    y: int,
    module_size: int,
    quality: int,
    orientation: str,
    text_offset_y: int,
    eac_offset_x: int,
    eac_offset_y: int,
) -> bytes:
    if orientation not in {"N", "R", "I", "B"}:
        raise ValueError("orientation must be N/R/I/B")

    text_gfa = image_to_zpl_gfa(text_img)
    eac_gfa = image_to_zpl_gfa(eac_img)

    return (
        b"^XA\r\n"
        + f"^FO{x},{y}\r\n".encode("ascii")
        + f"^BX{orientation},{module_size},{quality}\r\n".encode("ascii")
        + b"^FD" + payload + b"^FS\r\n"
        + f"^FO{x + eac_offset_x},{y + eac_offset_y}\r\n".encode("ascii")
        + eac_gfa + b"\r\n^FS\r\n"
        + f"^FO{x},{y + text_offset_y}\r\n".encode("ascii")
        + text_gfa + b"\r\n^FS\r\n"
        + b"^XZ\r\n"
    )


def build_blank_feed_command() -> bytes:
    return b"~PH"


def send_to_tcp(ip: str, port: int, data: bytes) -> None:
    with socket.create_connection((ip, port), timeout=10) as sock:
        sock.sendall(data)


def save_debug_file(filepath: str, data: bytes) -> None:
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def print_debug_info(
    excel_row: int,
    barcode_value,
    payload: bytes,
    gtin_lines: List[str],
    name_lines: List[str],
    ai21_lines: List[str],
    zpl: bytes,
    text_img: Image.Image,
    eac_img: Image.Image,
    eac_offset_x: int,
    eac_offset_y: int,
    text_offset_y: int,
) -> None:
    print(f"[DEBUG] Excel row: {excel_row}")
    print("[DEBUG] repr(barcode_value):")
    print(repr(barcode_value))
    print()
    print("[DEBUG] payload length:", len(payload))
    print("[DEBUG] payload hex:")
    print(payload.hex())
    print()
    print("[DEBUG] GS count:", payload.count(b"\x1d"))
    print("[DEBUG] payload visualized:")
    print(payload.replace(b"\x1d", b"<GS>"))
    print()
    print("[DEBUG] GTIN lines:")
    for i, line in enumerate(gtin_lines, start=1):
        print(f"  [{i}] {line}")
    print()
    print("[DEBUG] NAME lines:")
    for i, line in enumerate(name_lines, start=1):
        print(f"  [{i}] {line}")
    print()
    print("[DEBUG] AI21 lines:")
    for i, line in enumerate(ai21_lines, start=1):
        print(f"  [{i}] {line}")
    print()
    print("[DEBUG] text image size:", text_img.size)
    print("[DEBUG] eac image size:", eac_img.size)
    print("[DEBUG] eac offset:", eac_offset_x, eac_offset_y)
    print("[DEBUG] text offset y:", text_offset_y)
    print("[DEBUG] ZPL length:", len(zpl))
    print("[DEBUG] First 300 bytes of ZPL:")
    print(repr(zpl[:300]))
    print()


def parse_sheet_arg(value: str) -> Union[int, str]:
    try:
        return int(value)
    except ValueError:
        return value


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Print Data Matrix + text image + EAC + AI(21) via ZPL")
    parser.add_argument("--excel", required=True, help="Path to Excel file (.xlsx)")
    parser.add_argument("--sheet", default="0", help="Sheet index or name")
    parser.add_argument("--column", required=True, help="Column with Data Matrix")
    parser.add_argument("--gtin-column", required=True, help="Column with GTIN")
    parser.add_argument("--name-column", required=True, help="Column with item name/article")
    parser.add_argument("--row", type=int, default=0, help="0 = first data row after header")
    parser.add_argument("--all", action="store_true", help="Print all rows including blank separators")
    parser.add_argument("--ip", required=True, help="Printer IP")
    parser.add_argument("--port", type=int, default=9100, help="Printer TCP port")
    parser.add_argument("--x", type=int, default=100, help="Data Matrix X")
    parser.add_argument("--y", type=int, default=40, help="Data Matrix Y")
    parser.add_argument("--module-size", type=int, default=4, help="Data Matrix module size")
    parser.add_argument("--quality", type=int, default=200, help="^BX quality")
    parser.add_argument("--orientation", default="N", help="N/R/I/B")
    parser.add_argument("--text-offset-y", type=int, default=0, help="Text vertical offset below Data Matrix")
    parser.add_argument("--text-width-px", type=int, default=500, help="Text block width in pixels")
    parser.add_argument("--text-padding", type=int, default=4, help="Text image padding")
    parser.add_argument("--gtin-font-size", type=int, default=24, help="GTIN font size")
    parser.add_argument("--name-font-size", type=int, default=32, help="Name font size")
    parser.add_argument("--ai21-font-size", type=int, default=22, help="AI(21) font size")
    parser.add_argument("--line-gap", type=int, default=3, help="Line gap")
    parser.add_argument("--name-max-lines", type=int, default=2, help="Max lines for item name")
    parser.add_argument("--font-path", default=r"C:\Windows\Fonts\arial.ttf", help="Path to TTF font")
    parser.add_argument("--eac-logo-path", default="eac.png", help="Path to EAC logo file")
    parser.add_argument("--eac-height-px", type=int, default=72, help="EAC logo height in pixels")
    parser.add_argument("--eac-gap-px", type=int, default=16, help="Gap between code and EAC")
    parser.add_argument("--eac-offset-x", type=int, default=0, help="EAC X offset")
    parser.add_argument("--eac-offset-y", type=int, default=0, help="EAC Y offset")
    parser.add_argument("--save-zpl", help="Save final ZPL to file")
    parser.add_argument("--no-send", action="store_true", help="Do not send to printer, debug only")
    return parser


def process_blank_row(excel_row: int, args) -> None:
    cmd = build_blank_feed_command()
    print(f"[INFO] Excel row {excel_row}: blank separator row detected")
    print("[INFO] Sending one blank label feed via ~PH")

    if args.save_zpl:
        base = Path(args.save_zpl)
        out = base.with_name(f"{base.stem}_row_{excel_row}{base.suffix or '.zpl'}") if args.all else base
        save_debug_file(str(out), cmd)
        print(f"[INFO] Blank-feed command saved: {out}")

    if args.no_send:
        print("[INFO] Send disabled by --no-send")
        print()
        return

    print(f"[INFO] Sending blank label feed to {args.ip}:{args.port}")
    send_to_tcp(args.ip, args.port, cmd)
    print("[OK] Blank label feed sent")
    print()


def process_excel_row(ws, excel_row: int, args) -> None:
    barcode_value = get_cell_value_by_header(ws, excel_row, args.column)
    gtin_value = get_cell_value_by_header(ws, excel_row, args.gtin_column)
    name_value = get_cell_value_by_header(ws, excel_row, args.name_column)

    payload = build_payload(barcode_value)
    ai21_text = extract_ai_21(payload)

    gtin_lines, name_lines, ai21_lines = build_text_lines(
        gtin_text=normalize_text(gtin_value),
        name_text=normalize_text(name_value),
        ai21_text=ai21_text,
        font_path=args.font_path,
        gtin_font_size=args.gtin_font_size,
        name_font_size=args.name_font_size,
        ai21_font_size=args.ai21_font_size,
        text_width_px=args.text_width_px,
        name_max_lines=args.name_max_lines,
    )

    text_img = render_text_block_to_image(
        gtin_lines=gtin_lines,
        name_lines=name_lines,
        ai21_lines=ai21_lines,
        font_path=args.font_path,
        gtin_font_size=args.gtin_font_size,
        name_font_size=args.name_font_size,
        ai21_font_size=args.ai21_font_size,
        width_px=args.text_width_px,
        padding=args.text_padding,
        line_gap=args.line_gap,
    )

    eac_img = load_and_prepare_logo(args.eac_logo_path, args.eac_height_px)

    zpl = build_zpl(
        payload=payload,
        text_img=text_img,
        eac_img=eac_img,
        x=args.x,
        y=args.y,
        module_size=args.module_size,
        quality=args.quality,
        orientation=args.orientation,
        text_offset_y=args.text_offset_y,
        eac_offset_x=args.eac_offset_x,
        eac_offset_y=args.eac_offset_y,
    )

    print_debug_info(
        excel_row=excel_row,
        barcode_value=barcode_value,
        payload=payload,
        gtin_lines=gtin_lines,
        name_lines=name_lines,
        ai21_lines=ai21_lines,
        zpl=zpl,
        text_img=text_img,
        eac_img=eac_img,
        eac_offset_x=args.eac_offset_x,
        eac_offset_y=args.eac_offset_y,
        text_offset_y=args.text_offset_y,
    )

    if args.no_send:
        print("[INFO] Send disabled by --no-send")
        print()
        return

    send_to_tcp(args.ip, args.port, zpl)
    print("[OK] Label sent")
    print()


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    try:
        sheet_name = parse_sheet_arg(args.sheet)
        ws = get_worksheet(args.excel, sheet_name)

        if args.all:
            barcode_col_idx = find_column_index_by_header(ws, args.column)

            processed = 0
            empty_streak = 0
            excel_row = 2

            while True:
                cell_value = ws.cell(row=excel_row, column=barcode_col_idx).value

                if is_blank_separator(cell_value):
                    empty_streak += 1
                    if empty_streak >= 50:
                        break

                    process_blank_row(excel_row, args)
                    processed += 1
                    excel_row += 1
                    continue

                empty_streak = 0
                process_excel_row(ws, excel_row, args)
                processed += 1
                excel_row += 1

            print(f"[OK] Processed rows/feeds: {processed}")
            return 0

        excel_row = args.row + 2
        barcode_value = get_cell_value_by_header(ws, excel_row, args.column)

        if is_blank_separator(barcode_value):
            process_blank_row(excel_row, args)
        else:
            process_excel_row(ws, excel_row, args)

        return 0

    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
