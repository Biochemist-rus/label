import sys
import math
import socket
import argparse
from pathlib import Path
from typing import Union, List, Optional

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


def resource_path(relative_path: str) -> Path:
    """
    Возвращает путь к ресурсу:
    - при обычном запуске из .py
    - при запуске из собранного PyInstaller .exe
    """
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / relative_path
    return Path(__file__).resolve().parent / relative_path

def get_worksheet(excel_file: str, sheet_name: Union[int, str]):
    path = Path(excel_file)
    if not path.exists():
        raise FileNotFoundError(f"Excel-файл не найден: {excel_file}")

    wb = load_workbook(filename=excel_file, data_only=True)

    if isinstance(sheet_name, int):
        sheet_names = wb.sheetnames
        if sheet_name < 0 or sheet_name >= len(sheet_names):
            raise IndexError(f"sheet index {sheet_name} вне диапазона. Листы: {sheet_names}")
        ws = wb[sheet_names[sheet_name]]
    else:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Лист '{sheet_name}' не найден. Доступные листы: {wb.sheetnames}")
        ws = wb[sheet_name]

    return ws


def find_column_index_by_header(ws, column_name: str) -> int:
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        raise KeyError(f"Колонка '{column_name}' не найдена. Доступные заголовки: {headers}")
    return headers.index(column_name) + 1


def get_cell_value_by_header(ws, excel_row: int, column_name: str):
    col_idx = find_column_index_by_header(ws, column_name)
    return ws.cell(row=excel_row, column=col_idx).value


def normalize_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def build_payload(cell_value) -> bytes:
    if cell_value is None:
        raise ValueError("Ячейка Data Matrix пустая (None)")

    text = str(cell_value)
    if text == "":
        raise ValueError("Ячейка Data Matrix пустая (empty string)")

    text = text.replace("_x001D_", "\x1d")
    text = text.replace("_x001d_", "\x1d")

    try:
        payload = text.encode("latin-1")
    except UnicodeEncodeError as e:
        raise ValueError(
            "В значении Data Matrix есть символы вне latin-1. "
            "Их нужно отдельно нормализовать."
        ) from e

    return payload


def wrap_text_by_pixels(
    text: str,
    font: ImageFont.FreeTypeFont,
    max_width_px: int,
    max_lines: int,
    draw: ImageDraw.ImageDraw,
) -> List[str]:
    text = normalize_text(text)
    if not text:
        return []

    words = text.split(" ")
    lines: List[str] = []
    current = ""

    def text_width(s: str) -> int:
        if not s:
            return 0
        bbox = draw.textbbox((0, 0), s, font=font)
        return bbox[2] - bbox[0]

    for word in words:
        if text_width(word) <= max_width_px:
            candidate = word if not current else current + " " + word
            if text_width(candidate) <= max_width_px:
                current = candidate
            else:
                if current:
                    lines.append(current)
                    if len(lines) >= max_lines:
                        return lines
                current = word
        else:
            if current:
                lines.append(current)
                if len(lines) >= max_lines:
                    return lines
                current = ""

            chunk = ""
            for ch in word:
                candidate = chunk + ch
                if text_width(candidate) <= max_width_px:
                    chunk = candidate
                else:
                    if chunk:
                        lines.append(chunk)
                        if len(lines) >= max_lines:
                            return lines
                    chunk = ch
            if chunk:
                current = chunk

    if current and len(lines) < max_lines:
        lines.append(current)

    return lines[:max_lines]


def load_font(font_path: str, font_size: int) -> ImageFont.FreeTypeFont:
    path = Path(font_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл шрифта не найден: {font_path}")
    return ImageFont.truetype(str(path), font_size)

def load_and_prepare_logo(logo_path: str, logo_height_px: int) -> Image.Image:
    """
    Сначала пытается взять логотип по внешнему пути.
    Если файла там нет — ищет его как встроенный ресурс PyInstaller.
    """
    external = Path(logo_path)
    if external.exists():
        path = external
    else:
        path = resource_path(logo_path)

    if not path.exists():
        raise FileNotFoundError(f"Файл логотипа не найден: {path}")

    logo = Image.open(path).convert("RGBA")
    w, h = logo.size
    if h <= 0:
        raise ValueError("Некорректный размер логотипа")

    scale = logo_height_px / h
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    logo = logo.resize((new_w, new_h), Image.LANCZOS)

    bg = Image.new("RGBA", logo.size, (255, 255, 255, 255))
    bg.alpha_composite(logo)
    logo = bg.convert("L")
    logo = logo.point(lambda p: 0 if p < 200 else 255, mode="1")
    return logo


    path = Path(logo_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл логотипа не найден: {logo_path}")

    logo = Image.open(path).convert("RGBA")

    w, h = logo.size
    if h <= 0:
        raise ValueError("Некорректный размер логотипа")

    scale = logo_height_px / h
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))

    logo = logo.resize((new_w, new_h), Image.LANCZOS)

    # Переносим на белый фон, чтобы не было проблем с прозрачностью
    bg = Image.new("RGBA", logo.size, (255, 255, 255, 255))
    bg.alpha_composite(logo)
    logo = bg.convert("L")

    # В ч/б
    logo = logo.point(lambda p: 0 if p < 200 else 255, mode="1")
    return logo


def render_text_block_to_image(
    gtin_lines: List[str],
    name_lines: List[str],
    font_path: str,
    gtin_font_size: int,
    name_font_size: int,
    width_px: int,
    padding: int,
    line_gap: int,
    eac_logo_path: Optional[str] = None,
    eac_height_px: int = 64,
    eac_gap_px: int = 16,
) -> Image.Image:
    gtin_font = load_font(font_path, gtin_font_size)
    name_font = load_font(font_path, name_font_size)

    # Временный canvas для измерений
    tmp = Image.new("L", (width_px, 2000), 255)
    draw = ImageDraw.Draw(tmp)

    eac_logo = None
    eac_width = 0
    eac_height = 0

    if eac_logo_path:
        eac_logo = load_and_prepare_logo(eac_logo_path, eac_height_px)
        eac_width, eac_height = eac_logo.size

    text_x = padding
    if eac_logo is not None:
        text_x = padding + eac_width + eac_gap_px

    usable_text_width = max(40, width_px - text_x - padding)

    # На случай если строки уже не влезают после реального text_x
    gtin_lines = rewrap_lines(gtin_lines, gtin_font, usable_text_width, 1, draw)
    name_lines = rewrap_lines(name_lines, name_font, usable_text_width, len(name_lines) if name_lines else 0, draw)

    all_text_items = []
    for line in gtin_lines:
        bbox = draw.textbbox((0, 0), line, font=gtin_font)
        all_text_items.append((line, gtin_font, bbox[3] - bbox[1]))
    for line in name_lines:
        bbox = draw.textbbox((0, 0), line, font=name_font)
        all_text_items.append((line, name_font, bbox[3] - bbox[1]))

    if all_text_items:
        text_height = sum(h for _, _, h in all_text_items) + line_gap * (len(all_text_items) - 1)
    else:
        text_height = max(gtin_font_size, name_font_size)

    content_height = max(text_height, eac_height)
    height_px = padding * 2 + content_height

    img = Image.new("L", (width_px, height_px), 255)
    draw = ImageDraw.Draw(img)

    if eac_logo is not None:
        img.paste(eac_logo, (padding, padding))

    y = padding
    for line, font, h in all_text_items:
        draw.text((text_x, y), line, font=font, fill=0)
        y += h + line_gap

    bw = img.point(lambda p: 0 if p < 200 else 255, mode="1")
    return bw


def rewrap_lines(
    lines: List[str],
    font: ImageFont.FreeTypeFont,
    max_width_px: int,
    max_lines: int,
    draw: ImageDraw.ImageDraw,
) -> List[str]:
    if not lines:
        return []

    merged = " ".join(line.strip() for line in lines if line.strip())
    return wrap_text_by_pixels(
        merged,
        font=font,
        max_width_px=max_width_px,
        max_lines=max_lines if max_lines > 0 else 1,
        draw=draw,
    )


def image_to_zpl_gfa(img: Image.Image) -> bytes:
    if img.mode != "1":
        img = img.convert("1")

    width, height = img.size
    bytes_per_row = math.ceil(width / 8)
    total_bytes = bytes_per_row * height

    pixels = img.load()
    rows_hex = []

    for y in range(height):
        row_bytes = bytearray()
        current_byte = 0
        bit_count = 0

        for x in range(width):
            bit = 1 if pixels[x, y] == 0 else 0
            current_byte = (current_byte << 1) | bit
            bit_count += 1

            if bit_count == 8:
                row_bytes.append(current_byte)
                current_byte = 0
                bit_count = 0

        if bit_count > 0:
            current_byte = current_byte << (8 - bit_count)
            row_bytes.append(current_byte)

        rows_hex.append("".join(f"{b:02X}" for b in row_bytes))

    data = "".join(rows_hex)
    gfa = f"^GFA,{total_bytes},{total_bytes},{bytes_per_row},{data}"
    return gfa.encode("ascii")


def build_text_lines(
    gtin_text: str,
    name_text: str,
    font_path: str,
    gtin_font_size: int,
    name_font_size: int,
    text_width_px: int,
    name_max_lines: int,
    eac_logo_path: Optional[str] = None,
    eac_height_px: int = 64,
    eac_gap_px: int = 16,
) -> tuple[List[str], List[str]]:
    gtin_text = normalize_text(gtin_text)
    name_text = normalize_text(name_text)

    temp_img = Image.new("L", (text_width_px, 1000), 255)
    draw = ImageDraw.Draw(temp_img)

    gtin_font = load_font(font_path, gtin_font_size)
    name_font = load_font(font_path, name_font_size)

    eac_width = 0
    if eac_logo_path:
        eac_logo = load_and_prepare_logo(eac_logo_path, eac_height_px)
        eac_width = eac_logo.size[0] + eac_gap_px

    usable_text_width = max(40, text_width_px - 12 - eac_width)

    gtin_lines: List[str] = []
    name_lines: List[str] = []

    if gtin_text:
        gtin_lines = wrap_text_by_pixels(
            gtin_text,
            font=gtin_font,
            max_width_px=usable_text_width,
            max_lines=1,
            draw=draw,
        )

    if name_text:
        name_lines = wrap_text_by_pixels(
            name_text,
            font=name_font,
            max_width_px=usable_text_width,
            max_lines=name_max_lines,
            draw=draw,
        )

    return gtin_lines, name_lines


def build_zpl(
    payload: bytes,
    text_img: Image.Image,
    x: int,
    y: int,
    module_size: int,
    quality: int,
    orientation: str,
    text_offset_y: int,
) -> bytes:
    if orientation not in {"N", "R", "I", "B"}:
        raise ValueError("orientation должен быть N/R/I/B")

    text_gfa = image_to_zpl_gfa(text_img)

    zpl = (
        b"^XA\r\n"
        + f"^FO{x},{y}\r\n".encode("ascii")
        + f"^BX{orientation},{module_size},{quality}\r\n".encode("ascii")
        + b"^FD"
        + payload
        + b"^FS\r\n"
        + f"^FO{x},{y + text_offset_y}\r\n".encode("ascii")
        + text_gfa
        + b"\r\n^FS\r\n"
        + b"^XZ\r\n"
    )
    return zpl


def send_to_tcp(ip: str, port: int, data: bytes) -> None:
    with socket.create_connection((ip, port), timeout=10) as sock:
        sock.sendall(data)


def save_debug_file(filepath: str, data: bytes) -> None:
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def save_debug_image(filepath: str, img: Image.Image) -> None:
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path)


def print_debug_info(
    excel_row: int,
    barcode_value,
    payload: bytes,
    gtin_lines: List[str],
    name_lines: List[str],
    zpl: bytes,
    text_img: Image.Image,
) -> None:
    print(f"[DEBUG] Excel row: {excel_row}")
    print("[DEBUG] repr(barcode_value):")
    print(repr(barcode_value))
    print()

    print("[DEBUG] payload length:", len(payload))
    print("[DEBUG] payload hex:")
    print(payload.hex())
    print()

    print("[DEBUG] GS count:", payload.count(b"\x1d"))
    print("[DEBUG] payload visualized:")
    print(payload.replace(b"\x1d", b"<GS>"))
    print()

    print("[DEBUG] GTIN lines:")
    for i, line in enumerate(gtin_lines, start=1):
        print(f"  [{i}] {line}")
    print()

    print("[DEBUG] NAME lines:")
    for i, line in enumerate(name_lines, start=1):
        print(f"  [{i}] {line}")
    print()

    print("[DEBUG] text image size:", text_img.size)
    print("[DEBUG] ZPL length:", len(zpl))
    print("[DEBUG] First 300 bytes of ZPL:")
    print(repr(zpl[:300]))
    print()


def parse_sheet_arg(value: str) -> Union[int, str]:
    try:
        return int(value)
    except ValueError:
        return value


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Печать Data Matrix + текста картинкой + EAC через ZPL"
    )

    parser.add_argument("--excel", required=True, help="Путь к Excel-файлу .xlsx")
    parser.add_argument("--sheet", default="0", help="Лист: индекс или имя")

    parser.add_argument("--column", required=True, help="Колонка с Data Matrix")
    parser.add_argument("--gtin-column", required=True, help="Колонка с GTIN")
    parser.add_argument("--name-column", required=True, help="Колонка с названием/артикулом")

    parser.add_argument("--row", type=int, default=0, help="0 = первая строка после заголовка")
    parser.add_argument("--all", action="store_true", help="Печатать все непустые строки")

    parser.add_argument("--ip", required=True, help="IP принтера")
    parser.add_argument("--port", type=int, default=9100, help="TCP-порт")

    parser.add_argument("--x", type=int, default=100, help="X кода")
    parser.add_argument("--y", type=int, default=40, help="Y кода")
    parser.add_argument("--module-size", type=int, default=4, help="Размер модуля Data Matrix")
    parser.add_argument("--quality", type=int, default=200, help="Параметр ^BX")
    parser.add_argument("--orientation", default="N", help="N/R/I/B")

    parser.add_argument("--text-offset-y", type=int, default=155, help="Отступ текста вниз от кода")
    parser.add_argument("--text-width-px", type=int, default=420, help="Ширина текстового блока в пикселях")
    parser.add_argument("--text-padding", type=int, default=4, help="Внутренний отступ в текстовой картинке")
    parser.add_argument("--gtin-font-size", type=int, default=22, help="Размер шрифта GTIN")
    parser.add_argument("--name-font-size", type=int, default=28, help="Размер шрифта названия")
    parser.add_argument("--line-gap", type=int, default=3, help="Межстрочный интервал")
    parser.add_argument("--name-max-lines", type=int, default=2, help="Максимум строк для названия")

    parser.add_argument(
        "--font-path",
        default=r"C:\Windows\Fonts\arial.ttf",
        help="Путь к TTF-шрифту с кириллицей",
    )

    parser.add_argument(
        "--eac-logo-path",
        default="eac.png",
        help="Путь к файлу логотипа EAC",
    )
    parser.add_argument("--eac-height-px", type=int, default=72, help="Высота логотипа EAC в пикселях")
    parser.add_argument("--eac-gap-px", type=int, default=16, help="Отступ между EAC и текстом")

    parser.add_argument("--save-zpl", help="Сохранить ZPL в файл")
    parser.add_argument("--save-text-image", help="Сохранить текстовую картинку в PNG")
    parser.add_argument("--no-send", action="store_true", help="Только отладка, не печатать")

    return parser


def process_excel_row(ws, excel_row: int, args) -> None:
    barcode_value = get_cell_value_by_header(ws, excel_row, args.column)
    gtin_value = get_cell_value_by_header(ws, excel_row, args.gtin_column)
    name_value = get_cell_value_by_header(ws, excel_row, args.name_column)

    payload = build_payload(barcode_value)

    gtin_lines, name_lines = build_text_lines(
        gtin_text=normalize_text(gtin_value),
        name_text=normalize_text(name_value),
        font_path=args.font_path,
        gtin_font_size=args.gtin_font_size,
        name_font_size=args.name_font_size,
        text_width_px=args.text_width_px,
        name_max_lines=args.name_max_lines,
        eac_logo_path=args.eac_logo_path,
        eac_height_px=args.eac_height_px,
        eac_gap_px=args.eac_gap_px,
    )

    text_img = render_text_block_to_image(
        gtin_lines=gtin_lines,
        name_lines=name_lines,
        font_path=args.font_path,
        gtin_font_size=args.gtin_font_size,
        name_font_size=args.name_font_size,
        width_px=args.text_width_px,
        padding=args.text_padding,
        line_gap=args.line_gap,
        eac_logo_path=args.eac_logo_path,
        eac_height_px=args.eac_height_px,
        eac_gap_px=args.eac_gap_px,
    )

    zpl = build_zpl(
        payload=payload,
        text_img=text_img,
        x=args.x,
        y=args.y,
        module_size=args.module_size,
        quality=args.quality,
        orientation=args.orientation,
        text_offset_y=args.text_offset_y,
    )

    print_debug_info(
        excel_row=excel_row,
        barcode_value=barcode_value,
        payload=payload,
        gtin_lines=gtin_lines,
        name_lines=name_lines,
        zpl=zpl,
        text_img=text_img,
    )

    if args.save_text_image:
        base = Path(args.save_text_image)
        out = base.with_name(f"{base.stem}_row_{excel_row}{base.suffix or '.png'}") if args.all else base
        save_debug_image(str(out), text_img)
        print(f"[INFO] Текстовая картинка сохранена: {out}")

    if args.save_zpl:
        base = Path(args.save_zpl)
        out = base.with_name(f"{base.stem}_row_{excel_row}{base.suffix or '.zpl'}") if args.all else base
        save_debug_file(str(out), zpl)
        print(f"[INFO] ZPL сохранён: {out}")

    if args.no_send:
        print("[INFO] Отправка отключена флагом --no-send")
        print()
        return

    print(f"[INFO] Отправка по TCP: {args.ip}:{args.port}")
    send_to_tcp(args.ip, args.port, zpl)
    print("[OK] Команда отправлена на принтер")
    print()


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    try:
        sheet_name = parse_sheet_arg(args.sheet)
        ws = get_worksheet(args.excel, sheet_name)

        if args.all:
            barcode_col_idx = find_column_index_by_header(ws, args.column)

            printed = 0
            empty_streak = 0
            excel_row = 2

            while True:
                cell_value = ws.cell(row=excel_row, column=barcode_col_idx).value

                if cell_value is None:
                    empty_streak += 1
                    if empty_streak >= 50:
                        break
                    excel_row += 1
                    continue

                empty_streak = 0

                if str(cell_value).strip() == "":
                    excel_row += 1
                    continue

                process_excel_row(ws, excel_row, args)
                printed += 1
                excel_row += 1

            print(f"[OK] Обработано строк: {printed}")
            return 0

        excel_row = args.row + 2
        process_excel_row(ws, excel_row, args)
        return 0

    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())