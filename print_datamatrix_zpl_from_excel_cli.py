import sys
import socket
import argparse
from pathlib import Path
from typing import Union, List

from openpyxl import load_workbook


def get_worksheet(excel_file: str, sheet_name: Union[int, str]):
    path = Path(excel_file)
    if not path.exists():
        raise FileNotFoundError(f"Excel-файл не найден: {excel_file}")

    wb = load_workbook(filename=excel_file, data_only=True)

    if isinstance(sheet_name, int):
        sheet_names = wb.sheetnames
        if sheet_name < 0 or sheet_name >= len(sheet_names):
            raise IndexError(
                f"sheet index {sheet_name} вне диапазона. Листы: {sheet_names}"
            )
        ws = wb[sheet_names[sheet_name]]
    else:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Лист '{sheet_name}' не найден. Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

    return ws


def find_column_index_by_header(ws, column_name: str) -> int:
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        raise KeyError(
            f"Колонка '{column_name}' не найдена. Доступные заголовки: {headers}"
        )
    return headers.index(column_name) + 1


def read_excel_cell_value(
    excel_file: str,
    sheet_name: Union[int, str],
    column_name: str,
    row_index: int,
):
    ws = get_worksheet(excel_file, sheet_name)
    col_idx = find_column_index_by_header(ws, column_name)
    excel_row = row_index + 2
    return ws.cell(row=excel_row, column=col_idx).value


def build_payload(cell_value) -> bytes:
    if cell_value is None:
        raise ValueError("Ячейка пустая (None)")

    text = str(cell_value)
    if text == "":
        raise ValueError("Ячейка пустая (empty string)")

    text = text.replace("_x001D_", "\x1d")
    text = text.replace("_x001d_", "\x1d")

    try:
        payload = text.encode("latin-1")
    except UnicodeEncodeError as e:
        raise ValueError(
            "В значении есть символы вне latin-1. "
            "Такие данные нужно отдельно нормализовать."
        ) from e

    return payload


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = " ".join(text.split())
    return text


def escape_zpl_text(text: str) -> str:
    """
    Для ^FD опаснее всего служебные управляющие последовательности ZPL.
    Простая и практичная защита:
    - убираем переводы строк
    - заменяем ^ и ~
    """
    text = text.replace("\r", " ").replace("\n", " ")
    text = text.replace("^", " ")
    text = text.replace("~", " ")
    return text


def wrap_text(text: str, max_chars: int, max_lines: int) -> List[str]:
    """
    Простой перенос по словам.
    Если слово длиннее лимита — режем его.
    """
    text = normalize_text(text)
    if not text:
        return []

    words = text.split(" ")
    lines: List[str] = []
    current = ""

    for word in words:
        while len(word) > max_chars:
            if current:
                lines.append(current)
                current = ""
                if len(lines) >= max_lines:
                    return lines
            lines.append(word[:max_chars])
            word = word[max_chars:]
            if len(lines) >= max_lines:
                return lines

        candidate = word if not current else current + " " + word
        if len(candidate) <= max_chars:
            current = candidate
        else:
            lines.append(current)
            if len(lines) >= max_lines:
                return lines
            current = word

    if current and len(lines) < max_lines:
        lines.append(current)

    return lines[:max_lines]


def truncate_with_ellipsis(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    if max_chars <= 1:
        return text[:max_chars]
    return text[: max_chars - 1] + "…"


def build_text_lines(
    gtin_text: str,
    name_text: str,
    gtin_max_chars: int,
    name_max_chars: int,
    name_max_lines: int,
) -> List[str]:
    lines: List[str] = []

    gtin_text = escape_zpl_text(normalize_text(gtin_text))
    name_text = escape_zpl_text(normalize_text(name_text))

    if gtin_text:
        lines.append(truncate_with_ellipsis(gtin_text, gtin_max_chars))

    if name_text:
        wrapped = wrap_text(name_text, max_chars=name_max_chars, max_lines=name_max_lines)
        lines.extend(wrapped)

    return lines


def build_zpl(
    payload: bytes,
    text_lines: List[str],
    x: int,
    y: int,
    module_size: int,
    quality: int,
    orientation: str = "N",
    text_offset_y: int = 220,
    text_line_gap: int = 34,
    font_height: int = 28,
    font_width: int = 28,
) -> bytes:
    if orientation not in {"N", "R", "I", "B"}:
        raise ValueError("orientation должен быть N/R/I/B")

    zpl = (
        b"^XA\r\n"
        + f"^FO{x},{y}\r\n".encode("ascii")
        + f"^BX{orientation},{module_size},{quality}\r\n".encode("ascii")
        + b"^FD"
        + payload
        + b"^FS\r\n"
    )

    current_y = y + text_offset_y

    for line in text_lines:
        safe_line = escape_zpl_text(line)
        zpl += (
            f"^FO{x},{current_y}\r\n".encode("ascii")
            + f"^A0N,{font_height},{font_width}\r\n".encode("ascii")
            + f"^FD{safe_line}^FS\r\n".encode("utf-8")
        )
        current_y += text_line_gap

    zpl += b"^XZ\r\n"
    return zpl


def send_to_tcp(ip: str, port: int, data: bytes) -> None:
    with socket.create_connection((ip, port), timeout=5) as sock:
        sock.sendall(data)


def save_debug_file(filepath: str, data: bytes) -> None:
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def print_debug_info(
    excel_row: int,
    barcode_value,
    payload: bytes,
    text_lines: List[str],
    zpl: bytes,
) -> None:
    print(f"[DEBUG] Excel row: {excel_row}")
    print("[DEBUG] repr(barcode_value):")
    print(repr(barcode_value))
    print()

    print("[DEBUG] payload length:", len(payload))
    print("[DEBUG] payload hex:")
    print(payload.hex())
    print()

    gs_count = payload.count(b"\x1d")
    print("[DEBUG] GS count:", gs_count)

    printable = payload.replace(b"\x1d", b"<GS>")
    print("[DEBUG] payload visualized:")
    print(printable)
    print()

    print("[DEBUG] text lines:")
    for idx, line in enumerate(text_lines, start=1):
        print(f"  [{idx}] {line}")
    print()

    print("[DEBUG] ZPL length:", len(zpl))
    print("[DEBUG] First 300 bytes of ZPL:")
    print(repr(zpl[:300]))
    print()


def parse_sheet_arg(value: str) -> Union[int, str]:
    try:
        return int(value)
    except ValueError:
        return value


def get_cell_value_by_header(ws, excel_row: int, column_name: str):
    col_idx = find_column_index_by_header(ws, column_name)
    return ws.cell(row=excel_row, column=col_idx).value


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Печать Data Matrix + 2 текстовые строки из Excel через ZPL"
    )

    parser.add_argument("--excel", required=True, help="Путь к Excel-файлу .xlsx")
    parser.add_argument(
        "--sheet",
        default="0",
        help="Лист: индекс (например 0) или имя листа",
    )

    parser.add_argument(
        "--column",
        required=True,
        help="Колонка с данными Data Matrix",
    )
    parser.add_argument(
        "--gtin-column",
        required=True,
        help="Колонка с GTIN",
    )
    parser.add_argument(
        "--name-column",
        required=True,
        help="Колонка с названием / артикулом",
    )

    parser.add_argument(
        "--row",
        type=int,
        default=0,
        help="Индекс строки данных, 0 = первая строка после заголовка",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Печатать все непустые строки из колонки Data Matrix",
    )

    parser.add_argument("--ip", required=True, help="IP принтера")
    parser.add_argument("--port", type=int, default=9100, help="TCP-порт принтера")

    parser.add_argument("--x", type=int, default=110, help="X-координата Data Matrix")
    parser.add_argument("--y", type=int, default=40, help="Y-координата Data Matrix")

    parser.add_argument(
        "--module-size",
        type=int,
        default=4,
        help="Размер модуля Data Matrix",
    )
    parser.add_argument(
        "--quality",
        type=int,
        default=200,
        help="Параметр ^BX",
    )
    parser.add_argument(
        "--orientation",
        default="N",
        help="Ориентация: N, R, I, B",
    )

    parser.add_argument(
        "--text-offset-y",
        type=int,
        default=220,
        help="Отступ текста вниз от верхней точки Data Matrix",
    )
    parser.add_argument(
        "--text-line-gap",
        type=int,
        default=34,
        help="Вертикальный шаг между строками текста",
    )
    parser.add_argument(
        "--font-height",
        type=int,
        default=28,
        help="Высота шрифта",
    )
    parser.add_argument(
        "--font-width",
        type=int,
        default=28,
        help="Ширина шрифта",
    )

    parser.add_argument(
        "--gtin-max-chars",
        type=int,
        default=24,
        help="Максимум символов в строке GTIN",
    )
    parser.add_argument(
        "--name-max-chars",
        type=int,
        default=28,
        help="Максимум символов в строке названия",
    )
    parser.add_argument(
        "--name-max-lines",
        type=int,
        default=2,
        help="Максимум строк для названия",
    )

    parser.add_argument(
        "--save-zpl",
        help="Сохранить итоговый ZPL в файл для отладки",
    )
    parser.add_argument(
        "--no-send",
        action="store_true",
        help="Не отправлять на принтер, только показать отладку и при необходимости сохранить ZPL",
    )

    return parser


def process_excel_row(ws, excel_row: int, args) -> None:
    barcode_value = get_cell_value_by_header(ws, excel_row, args.column)
    gtin_value = get_cell_value_by_header(ws, excel_row, args.gtin_column)
    name_value = get_cell_value_by_header(ws, excel_row, args.name_column)

    payload = build_payload(barcode_value)

    text_lines = build_text_lines(
        gtin_text=normalize_text(gtin_value),
        name_text=normalize_text(name_value),
        gtin_max_chars=args.gtin_max_chars,
        name_max_chars=args.name_max_chars,
        name_max_lines=args.name_max_lines,
    )

    zpl = build_zpl(
        payload=payload,
        text_lines=text_lines,
        x=args.x,
        y=args.y,
        module_size=args.module_size,
        quality=args.quality,
        orientation=args.orientation,
        text_offset_y=args.text_offset_y,
        text_line_gap=args.text_line_gap,
        font_height=args.font_height,
        font_width=args.font_width,
    )

    print_debug_info(
        excel_row=excel_row,
        barcode_value=barcode_value,
        payload=payload,
        text_lines=text_lines,
        zpl=zpl,
    )

    if args.save_zpl:
        if args.all:
            save_path = Path(args.save_zpl)
            stem = save_path.stem
            suffix = save_path.suffix or ".zpl"
            numbered = save_path.with_name(f"{stem}_row_{excel_row}{suffix}")
            save_debug_file(str(numbered), zpl)
            print(f"[INFO] ZPL сохранён: {numbered}")
        else:
            save_debug_file(args.save_zpl, zpl)
            print(f"[INFO] ZPL сохранён: {args.save_zpl}")

    if args.no_send:
        print("[INFO] Отправка отключена флагом --no-send")
        print()
        return

    print(f"[INFO] Отправка по TCP: {args.ip}:{args.port}")
    send_to_tcp(args.ip, args.port, zpl)
    print("[OK] Команда отправлена на принтер")
    print()


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    try:
        sheet_name = parse_sheet_arg(args.sheet)
        ws = get_worksheet(args.excel, sheet_name)

        if args.all:
            barcode_col_idx = find_column_index_by_header(ws, args.column)

            printed = 0
            empty_streak = 0
            excel_row = 2

            while True:
                cell_value = ws.cell(row=excel_row, column=barcode_col_idx).value

                if cell_value is None:
                    empty_streak += 1
                    if empty_streak >= 50:
                        break
                    excel_row += 1
                    continue

                empty_streak = 0

                text = str(cell_value).strip()
                if text == "":
                    excel_row += 1
                    continue

                process_excel_row(ws, excel_row, args)
                printed += 1
                excel_row += 1

            print(f"[OK] Обработано строк: {printed}")
            return 0

        excel_row = args.row + 2
        process_excel_row(ws, excel_row, args)
        return 0

    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())