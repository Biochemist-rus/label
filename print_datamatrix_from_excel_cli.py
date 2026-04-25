import sys
import socket
import argparse
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


def get_worksheet(excel_file: str, sheet_name: Union[int, str]):
    path = Path(excel_file)
    if not path.exists():
        raise FileNotFoundError(f"Excel-файл не найден: {excel_file}")

    wb = load_workbook(filename=excel_file, data_only=True)

    if isinstance(sheet_name, int):
        sheet_names = wb.sheetnames
        if sheet_name < 0 or sheet_name >= len(sheet_names):
            raise IndexError(
                f"sheet index {sheet_name} вне диапазона. Листы: {sheet_names}"
            )
        ws = wb[sheet_names[sheet_name]]
    else:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Лист '{sheet_name}' не найден. Доступные листы: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

    return ws


def find_column_index_by_header(ws, column_name: str) -> int:
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        raise KeyError(
            f"Колонка '{column_name}' не найдена. Доступные заголовки: {headers}"
        )
    return headers.index(column_name) + 1


def read_excel_cell_value(
    excel_file: str,
    sheet_name: Union[int, str],
    column_name: str,
    row_index: int,
):
    ws = get_worksheet(excel_file, sheet_name)
    col_idx = find_column_index_by_header(ws, column_name)
    excel_row = row_index + 2
    return ws.cell(row=excel_row, column=col_idx).value


def build_payload(cell_value) -> bytes:
    if cell_value is None:
        raise ValueError("Ячейка пустая (None)")

    text = str(cell_value)
    if text == "":
        raise ValueError("Ячейка пустая (empty string)")

    # Excel escaping for GS control char
    text = text.replace("_x001D_", "\x1d")
    text = text.replace("_x001d_", "\x1d")

    try:
        payload = text.encode("latin-1")
    except UnicodeEncodeError as e:
        raise ValueError(
            "В значении есть символы вне latin-1. "
            "Такие данные нужно отдельно нормализовать."
        ) from e

    return payload


def build_zpl(
    payload: bytes,
    x: int,
    y: int,
    module_size: int,
    quality: int,
    orientation: str = "N",
) -> bytes:
    """
    Строит ZPL для Data Matrix.
    ^BXo,h,s,c,r,f,g,a
    Нам нужен базовый надёжный вариант.
    """
    if orientation not in {"N", "R", "I", "B"}:
        raise ValueError("orientation должен быть N/R/I/B")

    # Используем CRLF для совместимости
    return (
        b"^XA\r\n"
        + f"^FO{x},{y}\r\n".encode("ascii")
        + f"^BX{orientation},{module_size},{quality}\r\n".encode("ascii")
        + b"^FD"
        + payload
        + b"^FS\r\n"
        + b"^XZ\r\n"
    )


def send_to_tcp(ip: str, port: int, data: bytes) -> None:
    with socket.create_connection((ip, port), timeout=5) as sock:
        sock.sendall(data)


def save_debug_file(filepath: str, data: bytes) -> None:
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def print_debug_info(excel_row: int, cell_value, payload: bytes, zpl: bytes) -> None:
    print(f"[DEBUG] Excel row: {excel_row}")
    print("[DEBUG] repr(cell_value):")
    print(repr(cell_value))
    print()

    print("[DEBUG] payload length:", len(payload))
    print("[DEBUG] payload hex:")
    print(payload.hex())
    print()

    gs_count = payload.count(b"\x1d")
    print("[DEBUG] GS count:", gs_count)

    printable = payload.replace(b"\x1d", b"<GS>")
    print("[DEBUG] payload visualized:")
    print(printable)
    print()

    print("[DEBUG] ZPL length:", len(zpl))
    print("[DEBUG] First 200 bytes of ZPL:")
    print(repr(zpl[:200]))
    print()


def parse_sheet_arg(value: str) -> Union[int, str]:
    try:
        return int(value)
    except ValueError:
        return value


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Печать Data Matrix из Excel на сетевой принтер через ZPL"
    )

    parser.add_argument("--excel", required=True, help="Путь к Excel-файлу .xlsx")
    parser.add_argument(
        "--sheet",
        default="0",
        help="Лист: индекс (например 0) или имя листа",
    )
    parser.add_argument(
        "--column",
        required=True,
        help="Имя колонки с данными штрихкода",
    )
    parser.add_argument(
        "--row",
        type=int,
        default=0,
        help="Индекс строки данных, 0 = первая строка после заголовка",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Печатать все непустые строки из колонки",
    )

    parser.add_argument("--ip", required=True, help="IP принтера")
    parser.add_argument("--port", type=int, default=9100, help="TCP-порт принтера")

    parser.add_argument("--x", type=int, default=50, help="X-координата")
    parser.add_argument("--y", type=int, default=50, help="Y-координата")
    parser.add_argument(
        "--module-size",
        type=int,
        default=4,
        help="Размер модуля Data Matrix",
    )
    parser.add_argument(
        "--quality",
        type=int,
        default=200,
        help="Параметр качества/ECC для ^BX",
    )
    parser.add_argument(
        "--orientation",
        default="N",
        help="Ориентация: N, R, I, B",
    )

    parser.add_argument(
        "--save-zpl",
        help="Сохранить итоговый ZPL в файл для отладки",
    )
    parser.add_argument(
        "--no-send",
        action="store_true",
        help="Не отправлять на принтер, только показать отладку и при необходимости сохранить ZPL",
    )

    return parser


def process_one_value(cell_value, excel_row: int, args) -> None:
    payload = build_payload(cell_value)
    zpl = build_zpl(
        payload=payload,
        x=args.x,
        y=args.y,
        module_size=args.module_size,
        quality=args.quality,
        orientation=args.orientation,
    )

    print_debug_info(excel_row, cell_value, payload, zpl)

    if args.save_zpl:
        save_debug_file(args.save_zpl, zpl)
        print(f"[INFO] ZPL сохранён: {args.save_zpl}")

    if args.no_send:
        print("[INFO] Отправка отключена флагом --no-send")
        return

    print(f"[INFO] Отправка по TCP: {args.ip}:{args.port}")
    send_to_tcp(args.ip, args.port, zpl)
    print("[OK] Команда отправлена на принтер")
    print()


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    try:
        sheet_name = parse_sheet_arg(args.sheet)

        if args.all:
            ws = get_worksheet(args.excel, sheet_name)
            col_idx = find_column_index_by_header(ws, args.column)

            printed = 0
            empty_streak = 0
            row = 2

            while True:
                cell_value = ws.cell(row=row, column=col_idx).value

                if cell_value is None:
                    empty_streak += 1
                    if empty_streak >= 50:
                        break
                    row += 1
                    continue

                empty_streak = 0

                text = str(cell_value)
                if text == "":
                    row += 1
                    continue

                process_one_value(
                    cell_value=cell_value,
                    excel_row=row,
                    args=args,
                )
                printed += 1
                row += 1

            print(f"[OK] Обработано строк: {printed}")
            return 0

        cell_value = read_excel_cell_value(
            excel_file=args.excel,
            sheet_name=sheet_name,
            column_name=args.column,
            row_index=args.row,
        )

        process_one_value(
            cell_value=cell_value,
            excel_row=args.row + 2,
            args=args,
        )
        return 0

    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())